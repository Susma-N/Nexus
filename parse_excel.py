import openpyxl
import json
import sys

try:
    wb = openpyxl.load_workbook('aeration_blower.xlsx', data_only=False)
except Exception as e:
    print(f"Error loading workbook: {e}")
    sys.exit(1)

output = {}

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    output[sheetname] = []
    for row in sheet.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                cell_data = {
                    'coord': cell.coordinate,
                    'value': str(cell.value),
                }
                row_data.append(cell_data)
        if row_data:
            output[sheetname].append(row_data)

with open('aeration_blower_dump.json', 'w') as f:
    json.dump(output, f, indent=2)

print("Dumped aeration_blower.xlsx to aeration_blower_dump.json")
