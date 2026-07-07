import openpyxl
import json
import sys

def parse_excel(filename):
    try:
        wb_formulas = openpyxl.load_workbook(filename, data_only=False)
        wb_values = openpyxl.load_workbook(filename, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        sys.exit(1)

    output = {}

    for sheetname in wb_formulas.sheetnames:
        sheet_f = wb_formulas[sheetname]
        sheet_v = wb_values[sheetname]
        output[sheetname] = []
        for row_f, row_v in zip(sheet_f.iter_rows(values_only=False), sheet_v.iter_rows(values_only=False)):
            row_data = []
            for cell_f, cell_v in zip(row_f, row_v):
                if cell_f.value is not None or cell_v.value is not None:
                    cell_data = {
                        'coord': cell_f.coordinate,
                        'formula': str(cell_f.value) if str(cell_f.value).startswith('=') else None,
                        'value': str(cell_v.value),
                        'raw': str(cell_f.value)
                    }
                    row_data.append(cell_data)
            if row_data:
                output[sheetname].append(row_data)

    with open('ad_dump.json', 'w') as f:
        json.dump(output, f, indent=2)

    print(f"Dumped {filename} to ad_dump.json")

if __name__ == '__main__':
    parse_excel('anaerobic_digester.xlsx')
