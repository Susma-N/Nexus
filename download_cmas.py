import requests
import openpyxl

url = 'https://docs.google.com/spreadsheets/d/1nTcmNUbqjMEvig1L4tgl2w-hu0gnQzZgQTHNkPzwtTg/export?format=xlsx'
r = requests.get(url)
with open('cmas_denitrification.xlsx', 'wb') as f:
    f.write(r.content)

wb = openpyxl.load_workbook('cmas_denitrification.xlsx', data_only=False)
print("Sheets:", wb.sheetnames)

with open('cmas_dump.txt', 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"=== SHEET: {sheet_name} ===\n")
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate
                    val = str(cell.value).replace('\n', ' | ')
                    row_data.append(f"{coord}: {val}")
            if row_data:
                f.write(" | ".join(row_data) + "\n")
        f.write("\n")

print("Dumped cmas_dump.txt")
