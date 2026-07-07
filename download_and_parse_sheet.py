import requests
import openpyxl

url = 'https://docs.google.com/spreadsheets/d/1azVdapJstFYkFAEoUBnTiS2U9tNhTg6S/export?format=xlsx'
r = requests.get(url)
with open('target_sheet.xlsx', 'wb') as f:
    f.write(r.content)

wb = openpyxl.load_workbook('target_sheet.xlsx', data_only=False)
with open('target_sheet_dump.txt', 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"=== SHEET: {sheet_name} ===\n")
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate
                    val = str(cell.value).replace('\n', ' ')
                    row_data.append(f"{coord}: {val}")
            if row_data:
                f.write(" | ".join(row_data) + "\n")
        f.write("\n")

print("Dumped target_sheet_dump.txt")
